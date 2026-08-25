from collections import defaultdict
from datetime import datetime, timezone
from io import BytesIO
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased

from app.models import (
    Device,
    Employee,
    Message,
    MessageSenderType,
    Role,
    RoleAssignment,
    RoleCode,
    Witness,
    WitnessBan,
)


class ReportPermissionDeniedError(PermissionError):
    pass


class ReportService:
    ROLE_NAMES = {
        RoleCode.INSPECTOR: "Инспектор",
        RoleCode.ADMINISTRATOR: "Администратор",
        RoleCode.CHIEF: "Начальник",
    }

    @staticmethod
    async def _require_chief(db: AsyncSession, device_id: UUID) -> None:
        chief = await db.scalar(
            select(Employee.id)
            .join(Device, Device.id == Employee.device_id)
            .join(RoleAssignment, RoleAssignment.employee_id == Employee.id)
            .join(Role, Role.id == RoleAssignment.role_id)
            .where(
                Device.id == device_id,
                RoleAssignment.revoked_at.is_(None),
                Role.code == RoleCode.CHIEF,
            )
        )
        if chief is None:
            raise ReportPermissionDeniedError("Only chief can export reports")

    @staticmethod
    def _excel_datetime(value: datetime) -> datetime:
        if value.tzinfo is not None:
            return value.astimezone(timezone.utc).replace(tzinfo=None)
        return value

    @staticmethod
    def _finish_sheet(sheet) -> None:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
            sheet.column_dimensions[column[0].column_letter].width = width

    @classmethod
    def _build_workbook(
        cls,
        ban_rows: list[tuple],
        role_rows: list[tuple],
        message_rows: list[tuple],
    ) -> bytes:
        workbook = Workbook()
        bans = workbook.active
        bans.title = "Баны"
        bans.append(["Дата и время", "ID выдавшего бан", "ID получившего бан"])
        for issued_at, issuer_device_id, witness_device_id in ban_rows:
            bans.append(
                [
                    cls._excel_datetime(issued_at),
                    str(issuer_device_id),
                    str(witness_device_id),
                ]
            )
        cls._finish_sheet(bans)

        roles = workbook.create_sheet("Роли")
        roles.append(
            ["Дата и время", "ID совершившего действие", "Действие", "ID устройства"]
        )
        grouped: dict[UUID, list[tuple]] = defaultdict(list)
        for row in role_rows:
            grouped[row[0].employee_id].append(row)
        for assignments in grouped.values():
            assignments.sort(key=lambda row: row[0].assigned_at)
            for index, row in enumerate(assignments):
                assignment, role_code, target_device, assigned_by, revoked_by = row
                previous = assignments[index - 1] if index else None
                replaced = (
                    previous is not None
                    and previous[0].revoked_at is not None
                    and abs((assignment.assigned_at - previous[0].revoked_at).total_seconds()) <= 60
                )
                if replaced:
                    action = (
                        f"Роль «{cls.ROLE_NAMES[previous[1]]}» заменена "
                        f"на «{cls.ROLE_NAMES[role_code]}»"
                    )
                else:
                    action = f"Выдана роль «{cls.ROLE_NAMES[role_code]}»"
                roles.append(
                    [
                        cls._excel_datetime(assignment.assigned_at),
                        str(assigned_by) if assigned_by else "Система",
                        action,
                        str(target_device),
                    ]
                )

                next_row = assignments[index + 1] if index + 1 < len(assignments) else None
                followed_by_replacement = (
                    assignment.revoked_at is not None
                    and next_row is not None
                    and abs((next_row[0].assigned_at - assignment.revoked_at).total_seconds()) <= 60
                )
                if assignment.revoked_at is not None and not followed_by_replacement:
                    roles.append(
                        [
                            cls._excel_datetime(assignment.revoked_at),
                            str(revoked_by) if revoked_by else "Не зафиксировано",
                            f"Удалена роль «{cls.ROLE_NAMES[role_code]}»",
                            str(target_device),
                        ]
                    )
        cls._finish_sheet(roles)

        messages = workbook.create_sheet("Сообщения")
        messages.append(["ID сотрудника", "Количество сообщений"])
        for employee_device_id, count in message_rows:
            messages.append([str(employee_device_id), count])
        cls._finish_sheet(messages)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @classmethod
    async def generate_activity_report(
        cls, db: AsyncSession, requester_device_id: UUID
    ) -> bytes:
        await cls._require_chief(db, requester_device_id)

        issuer_device = aliased(Device)
        witness_device = aliased(Device)
        ban_rows = list(
            (
                await db.execute(
                    select(
                        WitnessBan.issued_at,
                        issuer_device.id,
                        witness_device.id,
                    )
                    .join(Employee, Employee.id == WitnessBan.issued_by_employee_id)
                    .join(issuer_device, issuer_device.id == Employee.device_id)
                    .join(Witness, Witness.id == WitnessBan.witness_id)
                    .join(witness_device, witness_device.id == Witness.device_id)
                    .order_by(WitnessBan.issued_at)
                )
            ).all()
        )

        target_device = aliased(Device)
        assigned_employee = aliased(Employee)
        assigned_device = aliased(Device)
        revoked_employee = aliased(Employee)
        revoked_device = aliased(Device)
        role_rows = list(
            (
                await db.execute(
                    select(
                        RoleAssignment,
                        Role.code,
                        target_device.id,
                        assigned_device.id,
                        revoked_device.id,
                    )
                    .join(Role, Role.id == RoleAssignment.role_id)
                    .join(Employee, Employee.id == RoleAssignment.employee_id)
                    .join(target_device, target_device.id == Employee.device_id)
                    .outerjoin(
                        assigned_employee,
                        assigned_employee.id == RoleAssignment.assigned_by_employee_id,
                    )
                    .outerjoin(assigned_device, assigned_device.id == assigned_employee.device_id)
                    .outerjoin(
                        revoked_employee,
                        revoked_employee.id == RoleAssignment.revoked_by_employee_id,
                    )
                    .outerjoin(revoked_device, revoked_device.id == revoked_employee.device_id)
                    .order_by(RoleAssignment.assigned_at)
                )
            ).all()
        )

        message_rows = list(
            (
                await db.execute(
                    select(Device.id, func.count(Message.id))
                    .join(Employee, Employee.device_id == Device.id)
                    .outerjoin(
                        Message,
                        (Message.sender_device_id == Device.id)
                        & (Message.sender_type == MessageSenderType.EMPLOYEE),
                    )
                    .group_by(Device.id)
                    .order_by(Device.id)
                )
            ).all()
        )
        return cls._build_workbook(ban_rows, role_rows, message_rows)
