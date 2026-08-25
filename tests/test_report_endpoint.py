from datetime import datetime, timezone
from io import BytesIO
from types import SimpleNamespace
from uuid import uuid4

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.core.database import get_db
from app.main import app
from app.models.role import RoleCode
from app.services.report_service import ReportPermissionDeniedError, ReportService


async def override_db():
    yield object()


def sample_report() -> bytes:
    now = datetime.now(timezone.utc)
    employee_id = uuid4()
    device_id = uuid4()
    assignment = SimpleNamespace(
        employee_id=employee_id,
        assigned_at=now,
        revoked_at=None,
    )
    return ReportService._build_workbook(
        [(now, uuid4(), uuid4())],
        [(assignment, RoleCode.CHIEF, device_id, None, None)],
        [(device_id, 3)],
    )


def test_report_contains_three_required_sheets() -> None:
    workbook = load_workbook(BytesIO(sample_report()))
    assert workbook.sheetnames == ["Баны", "Роли", "Сообщения"]
    assert workbook["Сообщения"]["B2"].value == 3


def test_chief_can_download_report(monkeypatch) -> None:
    device_id = uuid4()

    async def generate(_db, requester_id):
        assert requester_id == device_id
        return sample_report()

    monkeypatch.setattr(ReportService, "generate_activity_report", generate)
    app.dependency_overrides[get_db] = override_db
    with TestClient(app) as client:
        response = client.get(
            "/api/v1/reports/activity.xlsx",
            params={"requester_device_id": str(device_id)},
        )
    app.dependency_overrides.clear()

    assert response.status_code == 200
    assert response.content.startswith(b"PK")
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_non_chief_cannot_download_report(monkeypatch) -> None:
    async def generate(_db, _requester_id):
        raise ReportPermissionDeniedError("Only chief can export reports")

    monkeypatch.setattr(ReportService, "generate_activity_report", generate)
    app.dependency_overrides[get_db] = override_db
    with TestClient(app) as client:
        response = client.get(
            "/api/v1/reports/activity.xlsx",
            params={"requester_device_id": str(uuid4())},
        )
    app.dependency_overrides.clear()

    assert response.status_code == 403
